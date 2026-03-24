"""
订单处理平台 - Flask Web应用
整合 merge.py / cost.py / cost01.py 三个处理脚本
"""

import os
import re
import csv
import glob
import uuid
import shutil
import threading
import chardet
import pandas as pd

from pathlib import Path
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# =====================================================================
# 目录配置（与原脚本保持一致）
# =====================================================================
BASE_DIR = Path(__file__).parent

ORDERS_DIR     = BASE_DIR / "后台订单原表"
SHIPMENTS_DIR  = BASE_DIR / "后台货件表"
COST_DIR       = BASE_DIR / "成本"
CHANNEL_DIR    = BASE_DIR / "多渠道订单表"
FEE_DIR        = BASE_DIR / "费用表"
RESULT_DIR     = BASE_DIR / "results"

for d in [ORDERS_DIR, SHIPMENTS_DIR, COST_DIR, CHANNEL_DIR, FEE_DIR, RESULT_DIR]:
    d.mkdir(parents=True, exist_ok=True)

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}
MAX_CONTENT_LENGTH = 500 * 1024 * 1024
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

# 任务存储（内存 + 文件双写，重启后可恢复）
tasks: dict = {}

TASKS_DIR = BASE_DIR / "workspace"
TASKS_DIR.mkdir(exist_ok=True)

def _write_task_status(task: dict):
    """将task状态写入磁盘，防止重启丢失"""
    try:
        p = TASKS_DIR / f"{task['id']}.json"
        import json as _json
        p.write_text(_json.dumps({
            'id': task['id'],
            'type': task['type'],
            'status': task['status'],
            'progress': task['progress'],
            'message': task['message'],
            'result_files': task.get('result_files', {})
        }, ensure_ascii=False), encoding='utf-8')
    except Exception:
        pass

def _load_task_from_disk(task_id: str) -> dict | None:
    """从磁盘读取task，用于重启后恢复"""
    try:
        p = TASKS_DIR / f"{task_id}.json"
        if p.exists():
            import json as _json
            return _json.loads(p.read_text(encoding='utf-8'))
    except Exception:
        pass
    return None

# =====================================================================
# 工具函数
# =====================================================================

def safe_filename(filename: str) -> str:
    filename = filename.strip().replace('\\', '').replace('/', '')
    filename = re.sub(r'[^\w\.\u4e00-\u9fa5\-]', '', filename, flags=re.UNICODE)
    filename = ' '.join(filename.split())
    return filename or 'file'

def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def _log(task: dict, msg: str, pct: int = None):
    task['message'] = msg
    if pct is not None:
        task['progress'] = pct
    print(f"[{task['id'][:8]}] {msg}")
    _write_task_status(task)

# =====================================================================
# ① merge.py 逻辑
# =====================================================================

MONTH_MAP = {
    "Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
    "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12
}

HEADER_TRANSLATION_MAP = {
    "Datum/Uhrzeit":"date/time","Abrechnungsnummer":"settlement id",
    "Typ":"type","Bestellnummer":"order id","SKU":"sku",
    "Beschreibung":"description","Menge":"quantity","Marketplace":"marketplace",
    "Versand":"fulfillment","Ort der Bestellung":"order city",
    "Bundesland":"order state","Postleitzahl":"order postal",
    "Steuererhebungsmodell":"tax collection model","Umsätze":"product sales",
    "Produktumsatzsteuer":"product sales tax",
    "Gutschrift für Versandkosten":"shipping credits",
    "Steuer auf Versandgutschrift":"shipping credits tax",
    "Gutschrift für Geschenkverpackung":"gift wrap credits",
    "Steuer auf Geschenkverpackungsgutschriften":"gift wrap credits tax",
    "Rabatte aus Werbeaktionen":"promotional rebates",
    "Steuer auf Aktionsrabatte":"promotional rebates tax",
    "Einbehaltene Steuer auf Marketplace":"marketplace withheld tax",
    "Verkaufsgebühren":"selling fees",
    "Gebühren zu Versand durch Amazon":"fba fees",
    "Andere Transaktionsgebühren":"other transaction fees",
    "Andere":"other","Gesamt":"total"
}

def _detect_encoding(fp):
    with open(fp, 'rb') as f:
        raw = f.read(200000)
        r = chardet.detect(raw)
        return r['encoding'] or 'latin1'

def _detect_delimiter(fp, enc):
    with open(fp, 'r', encoding=enc, errors='ignore') as f:
        sample = f.read(5000)
        try:
            return csv.Sniffer().sniff(sample).delimiter
        except:
            return ','

def _detect_header_row(fp, enc, delim):
    with open(fp, 'r', encoding=enc, errors='ignore') as f:
        for i in range(10):
            line = f.readline()
            if not line:
                break
            if len(line.strip().split(delim)) > 5:
                return i
    return None

def _read_csv(fp):
    enc = _detect_encoding(fp)
    delim = _detect_delimiter(fp, enc)
    hrow = _detect_header_row(fp, enc, delim)
    if hrow is None:
        return pd.DataFrame()
    df = pd.read_csv(fp, encoding=enc, sep=delim, header=hrow,
                     engine='python', on_bad_lines='skip')
    df = df.dropna(axis=1, how='all')
    new_cols = {c: HEADER_TRANSLATION_MAP.get(c.strip(), c.strip()) for c in df.columns}
    df.rename(columns=new_cols, inplace=True)
    return df

def _order_month(filename):
    m = re.search(r'([A-Za-z]{3})Monthly(Transaction|UnifiedTransaction)', filename)
    return MONTH_MAP.get(m.group(1)) if m else None

def _shipment_month(filename):
    name = Path(filename).stem.strip()
    m = re.search(r'(\d+)月', name)
    if m:
        return int(m.group(1))
    m = re.search(r'(\d{6})', name)
    if m:
        return int(m.group(1)[2:4])
    return None

def _create_pivot(df):
    for col in ["quantity","fba fees","total"]:
        if col not in df.columns:
            return pd.DataFrame()
        def parse_num(x):
            if pd.isna(x): return 0
            if isinstance(x,(int,float)): return x
            s = str(x).replace(",",".")
            nums = re.findall(r"-?\d+\.?\d*", s)
            return sum(float(n) for n in nums) if nums else 0
        df[col] = df[col].apply(parse_num)
    pivot = pd.pivot_table(df, index="sku",
                           values=["quantity","fba fees","total"],
                           aggfunc='sum').reset_index()
    total = pd.DataFrame({"sku":["总计"],"quantity":[pivot["quantity"].sum()],
                          "fba fees":[pivot["fba fees"].sum()],"total":[pivot["total"].sum()]})
    return pd.concat([pivot, total], ignore_index=True)

def _color_total_row(writer, sheet_name):
    wb = writer.book
    ws = wb[sheet_name]
    fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")
    for col in range(1, ws.max_column + 1):
        ws.cell(row=ws.max_row, column=col).fill = fill

def run_merge(task: dict) -> Path:
    """执行 merge.py 逻辑，返回生成的 xlsx 路径"""
    _log(task, "📂 扫描订单文件...", 5)

    order_files = [(f, _order_month(f)) for f in os.listdir(ORDERS_DIR)
                   if f.lower().endswith('.csv')]
    order_files = [(f, m) for f, m in order_files if m is not None]
    order_files.sort(key=lambda x: x[1])

    if not order_files:
        raise ValueError("后台订单原表中未找到有效 CSV 文件（文件名需含月份标识）")

    months = [m for _, m in order_files]
    month_label = f"{min(months)}-{max(months)}" if len(set(months)) > 1 else str(min(months))
    output_file = CHANNEL_DIR / f"{month_label}月多渠道.xlsx"

    sheet_written = False
    total = len(order_files)

    with pd.ExcelWriter(str(output_file), engine='openpyxl') as writer:
        for idx, (file_name, month) in enumerate(order_files):
            _log(task, f"处理订单 {file_name}...", 5 + int(idx / total * 45))
            order_df = _read_csv(str(ORDERS_DIR / file_name))
            if order_df.empty:
                continue

            ship_files = [f for f in os.listdir(SHIPMENTS_DIR) if _shipment_month(f) == month]
            shipment_dict = {}
            if ship_files:
                sdf = _read_csv(str(SHIPMENTS_DIR / ship_files[0]))
                if not sdf.empty:
                    shipment_dict = dict(zip(sdf.iloc[:,0].astype(str), sdf.iloc[:,1].astype(str)))

            multi_rows, influencer_rows = [], []
            for _, row in order_df.iterrows():
                if len(row) < 4:
                    continue
                d_val = str(row.iloc[3]).strip()
                if d_val.startswith("S"):
                    lookup = str(shipment_dict.get(d_val, ""))
                    if lookup.startswith("P"):
                        multi_rows.append(row)
                    elif lookup.startswith("CON"):
                        influencer_rows.append(row)

            for rows, suffix in [(multi_rows, "多渠道"), (influencer_rows, "网红")]:
                df = pd.DataFrame(rows, columns=order_df.columns)
                if df.empty:
                    continue
                # 只写透析表，不写原始明细工作表
                pivot = _create_pivot(df.copy())
                if pivot.empty:
                    continue
                pname = f"{month}月{'透析' if suffix == '多渠道' else '网红透析'}"
                pivot.to_excel(writer, sheet_name=pname, index=False)
                _color_total_row(writer, pname)
                sheet_written = True

        if not sheet_written:
            pd.DataFrame({"提示":["没有符合条件的数据"]}).to_excel(writer, sheet_name="无数据", index=False)

    _log(task, f"✅ 多渠道订单表已生成: {output_file.name}", 50)
    return output_file

# =====================================================================
# ② cost.py 逻辑
# =====================================================================

FEE_SHEET_NAME = "发货成本"
COST_HEADER = [
    "日期","MSKU","ASIN","FNSKU","数量","产品成本","单款人工",
    "每票产品成本","物流费用","海外仓头程","FBA编号","货件名称","备注"
]
MONTH_FILL  = PatternFill(start_color="B5C6EA", end_color="B5C6EA", fill_type="solid")
TOTAL_FILL  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def _get_fee_path_and_months() -> tuple[Path, list]:
    files = list(CHANNEL_DIR.glob("*多渠道.xlsx"))
    if not files:
        raise FileNotFoundError("多渠道订单表未找到，请先执行合并步骤")
    filename = files[0].name
    month_part = filename.split("多渠道")[0].replace("月", "")
    if "-" in month_part:
        s, e = month_part.split("-")
        months = list(range(int(s), int(e) + 1))
    else:
        months = [int(month_part)]
    fee_path = FEE_DIR / f"{month_part}月费用表.xlsx"
    return fee_path, months

def _load_product_cost() -> pd.DataFrame:
    for f in COST_DIR.glob("*.xlsx"):
        if "fba" not in f.name.lower():
            df = pd.read_excel(str(f))
            df.rename(columns=lambda x: x.strip(), inplace=True)
            return df
    raise FileNotFoundError("未找到产品成本表（成本文件夹内非 fba 开头的 xlsx）")

def _load_fba_files() -> pd.DataFrame:
    files = list(COST_DIR.glob("fba*.xlsx")) + list(COST_DIR.glob("FBA*.xlsx"))
    if not files:
        raise FileNotFoundError("未找到 FBA 发货表（成本文件夹内 fba*.xlsx）")
    dfs = []
    for f in files:
        df = pd.read_excel(str(f))
        df.rename(columns=lambda x: x.strip(), inplace=True)
        dfs.append(df)
    combined = pd.concat(dfs, ignore_index=True)
    combined["创建时间"] = pd.to_datetime(combined["创建时间"])
    combined.sort_values("创建时间", inplace=True)
    combined["月份"] = combined["创建时间"].dt.month
    return combined

def run_cost(task: dict) -> Path:
    """执行 cost.py 逻辑"""
    _log(task, "📋 生成发货成本表...", 52)
    fee_path, month_list = _get_fee_path_and_months()
    product_cost_df = _load_product_cost()
    fba_df = _load_fba_files()

    if fee_path.exists():
        wb = load_workbook(str(fee_path))
        if FEE_SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(FEE_SHEET_NAME)
        else:
            ws = wb[FEE_SHEET_NAME]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = FEE_SHEET_NAME

    current_row = 1

    for month in month_list:
        month_df = fba_df[fba_df["月份"] == month].copy()
        month_df.sort_values("创建时间", inplace=True)

        # 表头（只写一次）
        if current_row == 1:
            for ci, cn in enumerate(COST_HEADER, 1):
                c = ws.cell(row=current_row, column=ci, value=cn)
                c.font = Font(bold=True)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                c.border = THIN_BORDER
            ws.freeze_panes = ws['A2']
            current_row += 1

        # 月份标题行（逐列填充，不合并，避免重复生成时 MergedCell 只读错误）
        for _ci in range(1, len(COST_HEADER) + 1):
            _mc = ws.cell(row=current_row, column=_ci, value=(f"{month}月" if _ci == 1 else None))
            _mc.fill = MONTH_FILL
            _mc.border = THIN_BORDER
            if _ci == 1:
                _mc.alignment = Alignment(horizontal='center', vertical='center')
                _mc.font = Font(bold=True)
        current_row += 1

        # 合并产品成本
        month_df = month_df.merge(
            product_cost_df[["ASIN","总成本","人工"]], on="ASIN", how="left"
        )
        month_df.rename(columns={"总成本":"产品成本","人工":"单款人工"}, inplace=True)
        month_df["产品成本"] = month_df["产品成本"].fillna(0)
        month_df["单款人工"] = month_df["单款人工"].fillna(0)
        month_df["每票产品成本"] = (month_df["产品成本"] + month_df["单款人工"]) * month_df["申报量"].fillna(0)

        start_row = current_row

        if month_df.empty:
            ws.cell(row=current_row, column=1, value="（无数据）")
            current_row += 1
            total_row = current_row
            ws.cell(total_row, COST_HEADER.index("单款人工")+1, "国内发货总计").fill = TOTAL_FILL
            ws.cell(total_row, COST_HEADER.index("数量")+1, 0).fill = TOTAL_FILL
            ws.cell(total_row, COST_HEADER.index("每票产品成本")+1, 0).fill = TOTAL_FILL
            current_row += 1
            continue

        for _, row in month_df.iterrows():
            for ci, cn in enumerate(COST_HEADER, 1):
                if cn == "日期":       val = row.get("创建时间")
                elif cn == "数量":     val = row.get("申报量")
                elif cn == "FBA编号":  val = row.get("货件编号")
                else:                  val = row.get(cn)
                c = ws.cell(row=current_row, column=ci, value=val)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = THIN_BORDER
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        # 合并重复 FBA编号/货件名称列（先缓存值，再合并，避免 MergedCell 只读错误）
        for cn in ["FBA编号","货件名称"]:
            ci = COST_HEADER.index(cn) + 1
            # 先把所有单元格的值缓存下来（合并后读取会返回 MergedCell 对象）
            col_vals = {r: ws.cell(r, ci).value for r in range(start_row, current_row)}
            sm = start_row
            for r in range(start_row + 1, current_row):
                if col_vals[r] == col_vals[r-1]:
                    continue
                if sm < r - 1:
                    ws.merge_cells(start_row=sm, start_column=ci, end_row=r-1, end_column=ci)
                    ws.cell(sm, ci).value = col_vals[sm]
                    ws.cell(sm, ci).alignment = Alignment(horizontal='center', vertical='center')
                sm = r
            if sm < current_row - 1:
                ws.merge_cells(start_row=sm, start_column=ci, end_row=current_row-1, end_column=ci)
                ws.cell(sm, ci).value = col_vals[sm]
                ws.cell(sm, ci).alignment = Alignment(horizontal='center', vertical='center')

        # 总计行
        total_row = current_row
        ws.cell(total_row, COST_HEADER.index("单款人工")+1, "国内发货总计").fill = TOTAL_FILL
        qty_col  = COST_HEADER.index("数量") + 1
        cost_col = COST_HEADER.index("每票产品成本") + 1
        ws.cell(total_row, qty_col,
                f"=SUM({get_column_letter(qty_col)}{start_row}:{get_column_letter(qty_col)}{current_row-1})"
                ).fill = TOTAL_FILL
        ws.cell(total_row, cost_col,
                f"=SUM({get_column_letter(cost_col)}{start_row}:{get_column_letter(cost_col)}{current_row-1})"
                ).fill = TOTAL_FILL
        current_row += 1

    # 自动列宽
    for i in range(1, len(COST_HEADER) + 1):
        max_len = max(
            (len(str(ws.cell(r, i).value)) for r in range(1, ws.max_row+1) if ws.cell(r, i).value),
            default=0
        ) + 2
        ws.column_dimensions[get_column_letter(i)].width = max_len

    wb.save(str(fee_path))
    _log(task, f"✅ 发货成本表已保存: {fee_path.name}", 75)
    return fee_path

# =====================================================================
# ③ cost01.py 逻辑
# =====================================================================

def _safe_float(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    try:
        return float(str(v).replace("￥","").replace(",","").strip())
    except:
        return 0.0

def _find_col(headers, keywords):
    for i, h in enumerate(headers):
        if not h: continue
        h_clean = str(h).replace(" ","").strip()
        for k in keywords:
            if k in h_clean:
                return i
    return None

def run_cost01(task: dict) -> Path:
    """执行 cost01.py 逻辑，追加"多渠道订单" Sheet 到费用表"""
    _log(task, "📊 生成多渠道订单费用明细...", 78)

    # 找文件
    fee_files = list(FEE_DIR.glob("*费用表.xlsx"))
    if not fee_files:
        raise FileNotFoundError("未找到费用表.xlsx，请先执行发货成本步骤")
    fee_file = fee_files[0]

    order_files = list(CHANNEL_DIR.glob("*.xlsx"))
    if not order_files:
        raise FileNotFoundError("未找到多渠道订单表")
    order_file = order_files[0]

    product_files = [f for f in COST_DIR.glob("*.xlsx") if "产品成本" in f.name]
    if not product_files:
        raise FileNotFoundError('未找到 产品成本.xlsx（文件名需含"产品成本"）')
    product_file = product_files[0]

    fee_wb    = load_workbook(str(fee_file))
    order_wb  = load_workbook(str(order_file))
    product_wb = load_workbook(str(product_file), data_only=True)
    product_ws = product_wb.active

    # 样式
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    month_fill2  = PatternFill("solid", fgColor="B5C6EA")
    red_fill    = PatternFill("solid", fgColor="F8CBAD")
    bold        = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")
    border      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # 产品成本字典
    ph = [c.value for c in product_ws[1]]
    sku_i  = _find_col(ph, ["SKU"])
    asin_i = _find_col(ph, ["ASIN"])
    cost_i = _find_col(ph, ["总成本","成本"])
    labor_i= _find_col(ph, ["人工"])
    ship_i = _find_col(ph, ["头程"])

    product_dict = {}
    for row in product_ws.iter_rows(min_row=2, values_only=True):
        if sku_i is None: continue
        sku = row[sku_i]
        if not sku: continue
        product_dict[str(sku)] = {
            "ASIN":  row[asin_i]  if asin_i  is not None else "",
            "cost":  _safe_float(row[cost_i])  if cost_i  is not None else 0.0,
            "labor": _safe_float(row[labor_i]) if labor_i is not None else 0.0,
            "ship":  _safe_float(row[ship_i])  if ship_i  is not None else 0.0,
        }

    # 获取月份
    months = sorted({
        int(m.group(1))
        for name in order_wb.sheetnames
        for m in [re.search(r'(\d+)月?透析$', name)]
        if m and "网红" not in name
    })

    target_sheet = "多渠道订单"
    if target_sheet in fee_wb.sheetnames:
        del fee_wb[target_sheet]
    ws = fee_wb.create_sheet(target_sheet)

    current_row = 1

    # ── 多渠道成本部分 ──
    for month in months:
        headers = ["SKU","ASIN","数量","单把成本","折扣","人工","头程","总成本"]
        for col, title in enumerate(headers, 1):
            c = ws.cell(current_row, col, title)
            c.font = bold; c.fill = header_fill
            c.alignment = center; c.border = border
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        c = ws.cell(current_row, 1, f"{month}月 多渠道")
        c.fill = month_fill2; c.font = bold; c.alignment = center
        current_row += 1

        sname = f"{month}月透析"
        if sname not in order_wb.sheetnames:
            continue
        month_ws = order_wb[sname]
        oh = [c.value for c in month_ws[1]]
        sku_i2 = _find_col(oh, ["sku","SKU"])
        qty_i2 = _find_col(oh, ["quantity","数量"])

        total_qty = total_cost_sum = 0
        for row in month_ws.iter_rows(min_row=2, values_only=True):
            if sku_i2 is None or qty_i2 is None: continue
            sku = str(row[sku_i2])
            qty = _safe_float(row[qty_i2])
            if sku not in product_dict: continue
            p = product_dict[sku]
            tc = (p["cost"] + p["labor"] + p["ship"]) * qty
            for col, val in enumerate([sku, p["ASIN"], qty, p["cost"], "", p["labor"], p["ship"], tc], 1):
                ws.cell(current_row, col, val).border = border
            total_qty += qty; total_cost_sum += tc
            current_row += 1

        for col, val in enumerate([None,None,total_qty,None,None,None,"总计",total_cost_sum], 1):
            c = ws.cell(current_row, col, val)
            if val is not None: c.font = bold
            c.border = border
        current_row += 2

    current_row += 3  # 空行分隔

    # ── 网红 FBA 配送费部分 ──
    for month in months:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
        c = ws.cell(current_row, 1, f"网红FBA配送费-{month}月")
        c.fill = red_fill; c.font = bold; c.alignment = center
        current_row += 1

        for col, title in enumerate(["SKU","ASIN","数量","多渠道配送费"], 1):
            c = ws.cell(current_row, col, title)
            c.font = bold; c.fill = header_fill
            c.alignment = center; c.border = border
        current_row += 1

        sname = f"{month}月网红透析"
        if sname not in order_wb.sheetnames:
            continue
        month_ws = order_wb[sname]
        oh = [c.value for c in month_ws[1]]
        sku_i2 = _find_col(oh, ["sku","SKU"])
        qty_i2 = _find_col(oh, ["quantity","数量"])
        fee_i2 = _find_col(oh, ["total","配送费"])

        total_fee = 0
        for row in month_ws.iter_rows(min_row=2, values_only=True):
            if sku_i2 is None: continue
            sku = str(row[sku_i2]).strip()
            if "总计" in sku: continue
            qty = _safe_float(row[qty_i2]) if qty_i2 is not None else 0
            fee = _safe_float(row[fee_i2]) if fee_i2 is not None else 0
            asin = product_dict.get(sku, {}).get("ASIN", "")
            for col, val in enumerate([sku, asin, qty, fee], 1):
                ws.cell(current_row, col, val).border = border
            total_fee += fee
            current_row += 1

        ws.cell(current_row, 3, "总计").font = bold
        ws.cell(current_row, 4, total_fee).font = bold
        for col in range(1, 5): ws.cell(current_row, col).border = border
        current_row += 2

    # 自动列宽
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[cl].width = max_len + 4

    fee_wb.save(str(fee_file))
    _log(task, f"✅ 多渠道订单费用明细已写入: {fee_file.name}", 95)
    return fee_file

# =====================================================================
# Flask 路由
# =====================================================================

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/session', methods=['POST'])
def create_session():
    """前端首次上传时调用，返回一个新的 task_id 供后续使用"""
    task_id = str(uuid.uuid4())
    return jsonify({'success': True, 'task_id': task_id})


@app.route('/api/upload', methods=['POST'])
def upload_files():
    """
    上传文件
    type 参数:
      - orders     → 后台订单原表/
      - shipments  → 后台货件表/
      - cost       → 成本/
    """
    file_type = request.form.get('type')
    if file_type not in ['orders', 'shipments', 'cost']:
        return jsonify({'success': False, 'error': '无效的文件类型'}), 400

    files = request.files.getlist('files')
    if not files:
        return jsonify({'success': False, 'error': '未提供文件'}), 400

    dir_map = {'orders': ORDERS_DIR, 'shipments': SHIPMENTS_DIR, 'cost': COST_DIR}
    dest_dir = dir_map[file_type]
    dest_dir.mkdir(parents=True, exist_ok=True)

    saved = []
    for f in files:
        if f and allowed_file(f.filename):
            fname = safe_filename(f.filename)
            f.save(str(dest_dir / fname))
            saved.append(fname)

    if not saved:
        return jsonify({'success': False, 'error': '没有有效文件'}), 400

    return jsonify({'success': True, 'files': saved, 'count': len(saved), 'type': file_type})


@app.route('/api/process', methods=['POST'])
def process_data():
    """
    启动处理任务
    type: "merge" | "cost" | "all"
    """
    data = request.get_json() or {}
    process_type = data.get('type', 'all')
    if process_type not in ['merge', 'cost', 'all']:
        return jsonify({'success': False, 'error': '无效的处理类型'}), 400

    # 校验文件
    if not list(ORDERS_DIR.glob('*.csv')):
        return jsonify({'success': False, 'error': '请先上传后台订单原表（CSV 格式）'}), 400
    if not list(SHIPMENTS_DIR.iterdir()):
        return jsonify({'success': False, 'error': '请先上传后台货件表'}), 400
    if process_type in ['cost', 'all']:
        cost_files = list(COST_DIR.glob('*.xlsx')) + list(COST_DIR.glob('*.xls'))
        if not cost_files:
            return jsonify({'success': False, 'error': '请先上传成本文件（xlsx 格式）'}), 400

    # 复用前端传来的 task_id（来自 /api/session），没有则新建
    task_id = data.get('task_id') or str(uuid.uuid4())
    task = {
        'id': task_id,
        'type': process_type,
        'status': 'processing',
        'progress': 0,
        'message': '准备中...',
        'created_at': datetime.now().isoformat(),
        'result_files': {}
    }
    tasks[task_id] = task

    thread = threading.Thread(target=_run_task, args=(task_id, process_type), daemon=True)
    thread.start()

    return jsonify({'success': True, 'task_id': task_id})


def _run_task(task_id: str, process_type: str):
    task = tasks[task_id]
    try:
        merge_out = None
        fee_out   = None

        if process_type in ['merge', 'all']:
            merge_out = run_merge(task)
            task['result_files']['merge'] = str(merge_out)

        if process_type in ['cost', 'all']:
            fee_out = run_cost(task)
            run_cost01(task)
            task['result_files']['cost'] = str(fee_out)

        task['status']   = 'completed'
        task['progress'] = 100
        task['message']  = '✅ 全部处理完成'
        _write_task_status(task)

    except Exception as e:
        import traceback
        task['status']  = 'failed'
        task['message'] = f'❌ 错误: {e}'
        task['progress'] = 0
        _write_task_status(task)
        print(traceback.format_exc())


@app.route('/api/task/<task_id>')
def get_task_status(task_id):
    if task_id not in tasks:
        # 尝试从磁盘恢复（重启/多worker场景）
        t = _load_task_from_disk(task_id)
        if t is None:
            return jsonify({'success': False, 'error': '任务不存在'}), 404
        tasks[task_id] = t
    t = tasks[task_id]
    return jsonify({
        'success': True,
        'task': {
            'id': t['id'], 'type': t['type'],
            'status': t['status'], 'progress': t['progress'],
            'message': t['message'],
            'result_files': list(t.get('result_files', {}).keys())
        }
    })


@app.route('/api/download/<task_id>')
def download_zip(task_id):
    """将本次任务所有结果文件打包成 ZIP 下载"""
    import zipfile, io, tempfile

    # 先查内存，再查磁盘
    if task_id not in tasks:
        t = _load_task_from_disk(task_id)
        if t is None:
            return jsonify({'success': False, 'error': '任务不存在'}), 404
        tasks[task_id] = t
    task = tasks[task_id]

    if task['status'] != 'completed':
        return jsonify({'success': False, 'error': '任务未完成'}), 400

    # 收集所有结果文件
    result_files = task.get('result_files', {})
    files_to_zip = []
    for key, fp in result_files.items():
        p = Path(fp)
        if p.exists():
            files_to_zip.append(p)

    if not files_to_zip:
        return jsonify({'success': False, 'error': '结果文件不存在'}), 404

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for p in files_to_zip:
            zf.write(p, p.name)
    buf.seek(0)

    return send_file(buf, as_attachment=True,
                     download_name='报表结果.zip',
                     mimetype='application/zip')


@app.route('/api/download/<task_id>/<file_type>')
def download_file(task_id, file_type):
    if task_id not in tasks:
        return jsonify({'success': False, 'error': '任务不存在'}), 404
    task = tasks[task_id]
    if task['status'] != 'completed':
        return jsonify({'success': False, 'error': '任务未完成'}), 400

    fp = task.get('result_files', {}).get(file_type)
    name_map = {'merge': '多渠道订单表.xlsx', 'cost': '费用表.xlsx'}

    if not fp or not Path(fp).exists():
        return jsonify({'success': False, 'error': '文件不存在'}), 404

    return send_file(fp, as_attachment=True,
                     download_name=name_map.get(file_type, 'result.xlsx'),
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/cleanup', methods=['POST'])
def cleanup():
    """清空上传文件（保留结果）"""
    try:
        for d in [ORDERS_DIR, SHIPMENTS_DIR, COST_DIR, CHANNEL_DIR, FEE_DIR]:
            if d.exists():
                shutil.rmtree(d)
                d.mkdir(parents=True, exist_ok=True)
        return jsonify({'success': True, 'message': '已清空所有上传文件'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/files')
def list_files():
    """列出各目录文件状态，方便前端展示"""
    def ls(d: Path):
        return [f.name for f in d.iterdir()] if d.exists() else []
    return jsonify({
        'orders':    ls(ORDERS_DIR),
        'shipments': ls(SHIPMENTS_DIR),
        'cost':      ls(COST_DIR),
        'results':   ls(CHANNEL_DIR) + ls(FEE_DIR)
    })


@app.errorhandler(404)
def not_found(_): return jsonify({'success': False, 'error': '页面不存在'}), 404

@app.errorhandler(500)
def server_error(_): return jsonify({'success': False, 'error': '服务器内部错误'}), 500


if __name__ == '__main__':
    print("=" * 60)
    print("订单处理平台启动")
    print(f"  后台订单原表: {ORDERS_DIR}")
    print(f"  后台货件表:   {SHIPMENTS_DIR}")
    print(f"  成本文件夹:   {COST_DIR}")
    print(f"  多渠道输出:   {CHANNEL_DIR}")
    print(f"  费用表输出:   {FEE_DIR}")
    print("服务地址: http://127.0.0.1:5000")
    print("=" * 60)
    app.run(host='127.0.0.1', port=5000, debug=True, threaded=True)
